from openpyxl import load_workbook
from datetime import datetime
import os
import time

def execute(sourceFilePath,templateFilePath,finalFileSavePath):

    def dealDateDuring(date_info):
        projectCountDate = date_info.split("---")
        projectCountStartDate = projectCountDate[0].replace("（", "")
        projectCountStartDate = int(projectCountStartDate)
        projectCountEndDate = projectCountDate[1].replace("）", "")
        projectCountEndDate = int(projectCountEndDate)

        # 本周统计开始时间
        projectCountStartDate = '{:%Y.%m.%d}'.format(
            datetime(int(str(projectCountStartDate)[0:4]), int(str(projectCountStartDate)[4:6]),
                     int(str(projectCountStartDate)[6:8])))
        # 本周统计结束时间
        projectCountEndDate = '{:%Y.%m.%d}'.format(
            datetime(int(str(projectCountEndDate)[0:4]), int(str(projectCountEndDate)[4:6]),
                     int(str(projectCountEndDate)[6:8])))

        dateDealResultArray = [projectCountStartDate, projectCountEndDate]

        return dateDealResultArray


    wb = load_workbook(sourceFilePath)

    # 获得表单名字
    sheetNames = wb.sheetnames

    sheet = wb.get_sheet_by_name(sheetNames[0])

    # 数据截止时间
    validDate = str(sheet["B1"].value).split("：")[1].split(" ")[0]

    # 总交易金额
    totalTradeAmount = sheet["D4"].value
    # totalTradeAmount = format(totalTradeAmount,",")

    # 总交易笔数
    totalTradeCount = sheet["D5"].value

    # 总用户数
    totalAccountCount = sheet["D6"].value
    # 总收益
    totalAccountIncome = sheet["D7"].value
    # 本周兑付数
    currentWeekProjectCount = str(sheet["C11"].value)

    # 本周兑付金额
    currentWeekProjectAmount = format(sheet["C12"].value, ",")

    dateInfo = str(sheet["D12"].value)

    # 统计时间信息数组
    dateResultArray = dealDateDuring(dateInfo)

    dateDuringStart = dateResultArray[0]
    dateDuringEnd = dateResultArray[1]

    early60 = str(sheet["C18"].value)
    late60 = str(sheet["C19"].value)
    late70 = str(sheet["C20"].value)
    late80 = str(sheet["C21"].value)
    late90 = str(sheet["C22"].value)

    product15 = str(sheet["J6"].value)
    product30 = str(sheet["J8"].value)
    product60 = str(sheet["J10"].value)
    product90 = str(sheet["J11"].value)
    product180 = str(sheet["J7"].value)
    product360 = str(sheet["J9"].value)

    productSmallerThan7 = str(sheet["J19"].value)
    product7to8 = str(sheet["J18"].value)
    product8to10 = str(sheet["J20"].value)
    productBiggerThan10 = str(sheet["J17"].value)

    # print(validDate)
    # print(totalTradeAmount)
    # print(totalTradeCount)
    # print(totalAccountCount)
    # print(totalAccountIncome)
    # print(currentWeekProjectCount)
    # print(currentWeekProjectAmount)
    # print(dateDuringStart)
    # print(dateDuringEnd)
    # print(early60)
    # print(late60)
    # print(late70)
    # print(late80)
    # print(late90)
    #
    # print(product15)
    # print(product30)
    # print(product60)
    # print(product90)
    # print(product180)
    # print(product360)
    #
    # print(productSmallerThan7)
    # print(product7to8)
    # print(product8to10)
    # print(productBiggerThan10)

    templateFile = open(templateFilePath, 'r', encoding="utf-8")
    targetLine = ""
    for line in templateFile:
        if line.find("{validDate}") != -1:
            line = line.replace("{validDate}", validDate)
        if line.find("{totalTradeAmount}") != -1:
            line = line.replace("{totalTradeAmount}", totalTradeAmount)
        if line.find("{totalTradeCount}") != -1:
            line = line.replace("{totalTradeCount}", totalTradeCount)
        if line.find("{totalAccountCount}") != -1:
            line = line.replace("{totalAccountCount}", totalAccountCount)
        if line.find("{totalAccountIncome}") != -1:
            line = line.replace("{totalAccountIncome}", totalAccountIncome)
        if line.find("{currentWeekProjectCount}") != -1:
            line = line.replace("{currentWeekProjectCount}", currentWeekProjectCount)
        if line.find("{currentWeekProjectAmount}") != -1:
            line = line.replace("{currentWeekProjectAmount}", currentWeekProjectAmount)
        if line.find("{dateDuringStart}-{dateDuringEnd}") != -1:
            line = line.replace("{dateDuringStart}-{dateDuringEnd}", dateDuringStart + "-" + dateDuringEnd)

        if line.find("{early60}") != -1:
            line = line.replace("{early60}", early60)
        if line.find("{late60}") != -1:
            line = line.replace("{late60}", late60)
        if line.find("{late70}") != -1:
            line = line.replace("{late70}", late70)
        if line.find("{late80}") != -1:
            line = line.replace("{late80}", late80)
        if line.find("{late90}") != -1:
            line = line.replace("{late90}", late90)

        if line.find("{product15}") != -1:
            line = line.replace("{product15}", product15)
        if line.find("{product30}") != -1:
            line = line.replace("{product30}", product30)
        if line.find("{product60}") != -1:
            line = line.replace("{product60}", product60)
        if line.find("{product90}") != -1:
            line = line.replace("{product90}", product90)
        if line.find("{product180}") != -1:
            line = line.replace("{product180}", product180)
        if line.find("{product360}") != -1:
            line = line.replace("{product360}", product360)

        if line.find("{productSmallerThan7}") != -1:
            line = line.replace("{productSmallerThan7}", productSmallerThan7)
        if line.find("{product7to8}") != -1:
            line = line.replace("{product7to8}", product7to8)
        if line.find("{product8to10}") != -1:
            line = line.replace("{product8to10}", product8to10)
        if line.find("{productBiggerThan10}") != -1:
            line = line.replace("{productBiggerThan10}", productBiggerThan10)

        targetLine += line

    # date = time.strftime("%Y-%m-%d", time.localtime(time.time()))
    #
    # os.path.join("aboutDatashow/target/", date)
    #
    # # 创建一个目录:
    # if not os.path.exists("aboutDatashow/target/" + date):
    #     os.mkdir("aboutDatashow/target/" + date)

    fileName = "index.html"
    targetDir = finalFileSavePath
    os.chdir(targetDir)
    fo = open(fileName, "w", encoding='utf8')
    fo.write(targetLine)
    fo.flush()
    fo.close()
    return True
